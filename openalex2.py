#!/usr/bin/env python3
import time
from pathlib import Path
import requests
from pprint import pprint
import utils
from tqdm import tqdm
from rapidfuzz import fuzz
from collections import Counter
from openpyxl import load_workbook

API_KEY = "bdDEtP2Jp4MhNyiG42Ckzv"
BASE_URL = "https://api.openalex.org/works/"
HEADERS = { "User-Agent": "zotero-citation-analyzer/1.0" }
STORAGE_DIR = Path.home().joinpath( ".zotero-cg" , "openalex" )
MAX_RETRIES = 10

# ---------- search helpers ----------

def reconstruct_abstract( inv_index ):
	if not inv_index: return ""
	positions = [ ( pos , word ) for word , poses in inv_index.items() for pos in poses ]
	positions.sort()
	return " ".join( w for _ , w in positions )

def make_haystack( meta ):
	title = meta.get( "title" ) or meta.get( "display_name" ) or ""
	abstract = reconstruct_abstract( meta.get( "abstract_inverted_index" ) )
	return ( title + " " + abstract ).lower()

def fuzzy_has( haystack , term , threshold=80 ):
	return fuzz.partial_ratio( term.lower() , haystack ) >= threshold

def any_of( *terms , threshold=80 ):
	return lambda h: any( fuzzy_has( h , t , threshold ) for t in terms )

def all_of( *terms , threshold=80 ):
	return lambda h: all( fuzzy_has( h , t , threshold ) for t in terms )

def none_of( *terms , threshold=80 ):
	return lambda h: not any( fuzzy_has( h , t , threshold ) for t in terms )

def combine_and( *predicates ):
	return lambda h: all( p( h ) for p in predicates )

def combine_or( *predicates ):
	return lambda h: any( p( h ) for p in predicates )

# ---------- row helpers ----------

def title_of( d ):
	return d.get( "title" ) or d.get( "display_name" ) or ""

def build_row( wid , meta , cite_count ):
	rd = meta.get( "doi" )
	clean_doi = utils.normalize_doi( rd ) if rd else None
	proxy_url = f"https://doi-org.ezproxy.libraries.wright.edu/{clean_doi}" if clean_doi else None
	doi_url   = f"https://doi.org/{clean_doi}" if clean_doi else None
	proxy = utils.Link( proxy_url , proxy_url ) if proxy_url else ""
	link  = utils.Link( doi_url , doi_url ) if doi_url else ""
	return [
		cite_count,
		title_of( meta ) or "(no metadata)",
		meta.get( "publication_year" ),
		proxy,
		clean_doi,
		link,
		meta.get( "cited_by_count" ),
		wid,
	]

HEADERS_ROW = [ "Cites" , "Title" , "Year" , "Proxy" , "DOI" , "Link" , "OA Cited-By" , "WID" ]

def safe_sheet_name( name ):
	safe = name
	for ch in ":/\\?*[]":
		safe = safe.replace( ch , " " )
	return safe.strip()[ :31 ]

# ---------- main class ----------

class OpenAlex:
	def __init__( self , options={} ):
		self.options = options
		self.api_key = options.get( "api_key" , API_KEY )
		self.base_url = options.get( "base_url" , BASE_URL )
		self.headers = options.get( "headers" , HEADERS )
		self.params = { "api_key": self.api_key }
		self.storage_dir = options.get( "storage_dir" , STORAGE_DIR )
		self.storage_dir.mkdir( parents=True , exist_ok=True )
		self.references_dir = self.storage_dir.joinpath( "references" )
		self.references_dir.mkdir( parents=True , exist_ok=True )
		self.problem_dois = set()

		# Populated by stats(); reused by add_search_sheets()
		self._index = []   # list of ( wid , meta , haystack , cite_count , included_in_missing )
		self._xlsx_path = self.storage_dir / "missing.xlsx"

	# ---------- api ----------

	def get_zotero_id( self , zotero_item ):
		_save_path = self.storage_dir.joinpath( f"{zotero_item['key']}.json" )
		return zotero_item.get( "key" )

	def api_get_doi( self , doi ):
		url = self.base_url + f"https://doi.org/{doi}"
		while True:
			r = requests.get( url , params=self.params , headers=self.headers )
			if r.status_code == 200:
				return r.json()
			elif r.status_code == 429:
				retry = int( r.headers.get( "Retry-After" , 5 ) )
				print( f"\nRate limited. Sleeping {retry}s" )
				time.sleep( retry )
			else:
				return None

	def api_search_title( self , title , per_page=10 ):
		url = self.base_url.rstrip( "/" )
		params = dict( self.params )
		params.update({
			"search": f'"{title}"',
			"per-page": per_page,
			"select": "id,doi,title,display_name,publication_year,cited_by_count,relevance_score,authorships"
		})
		while True:
			r = requests.get( url , params=params , headers=self.headers )
			if r.status_code == 200:
				return r.json().get( "results" , [] )
			elif r.status_code == 429:
				retry = int( r.headers.get( "Retry-After" , 5 ) )
				print( f"\nRate limited. Sleeping {retry}s" )
				time.sleep( retry )
			else:
				print( f"\nOpenAlex title search failed: {r.status_code} {r.url} {r.text[:500]}" )
				return []

	def api_get_id( self , open_alex_wid ):
		url = BASE_URL + open_alex_wid
		for attempt in range( MAX_RETRIES ):
			try:
				r = requests.get( url , params=self.params , headers=self.headers , timeout=30 )
				if r.status_code == 200:
					return r.json()
				elif r.status_code == 429:
					retry = int( r.headers.get( "Retry-After" , 5 ) )
					print( f"\nRate limited resolving refs. Sleeping {retry}s" )
					time.sleep( retry )
					continue
				else:
					return None
			except requests.exceptions.RequestException as e:
				wait = min( 2 ** attempt , 60 )
				print( f"\nNetwork error ({e}). Retry {attempt+1}/{MAX_RETRIES}. Sleeping {wait}s" )
				time.sleep( wait )
		print( f"\nFailed after {MAX_RETRIES} retries: {open_alex_wid}" )
		return None

	# ---------- cache ----------

	def update_cache( self ):
		self.zotero_snapshot = utils.zotero_simple_snapshot()
		self.zotero_snapshot_keys = self.zotero_snapshot.keys()
		for k , key in enumerate( tqdm( self.zotero_snapshot_keys , desc="Papers" , position=0 ) ):

			# 1.) Download Info for Papers in Zotero Library
			paper_doi = self.zotero_snapshot[ key ].get( "doi" )
			paper_title = self.zotero_snapshot[ key ].get( "title" )
			paper_title_normalized = utils.openalex_normalize_title( paper_title )
			z_id = self.zotero_snapshot[ key ].get( "id" )
			zotero_cached_fp = self.storage_dir.joinpath( f"{z_id}.json" )
			if not paper_doi:
				if zotero_cached_fp.exists():
					continue
				else:
					print( "searching title" , paper_title_normalized )
					search_results = self.api_search_title( paper_title_normalized )
					if len( search_results ) > 1:
						paper_dois = [ p.get( "doi" ) for p in search_results if p.get( "doi" ) is not None ]
						if len( paper_dois ) > 1:
							paper_doi = paper_dois[ 0 ]
						else:
							print( "still nothing" , self.zotero_snapshot[ key ] , search_results )
					utils.write_json( zotero_cached_fp , { "id": z_id , "doi": paper_doi , "title": paper_title } )
			if not paper_doi:
				continue
			paper_doi_normalized = utils.normalize_doi( paper_doi )
			paper_doi_b64 = utils.base64_encode( paper_doi_normalized )
			paper_cached_fp = self.storage_dir.joinpath( f"{paper_doi_b64}.json" )
			if paper_cached_fp.exists():
				continue
			paper_data = self.api_get_doi( paper_doi )
			utils.write_json( paper_cached_fp , paper_data )

			# 2.) Download all of its References
			referenced_works = paper_data.get( "referenced_works" )
			if not referenced_works:
				continue
			for i , item in enumerate( tqdm( referenced_works , desc="References" , position=1 , leave=False ) ):
				wid = item.split( "/" )[ -1 ]
				reference_cached_fp = self.references_dir.joinpath( f"{wid}.json" )
				if reference_cached_fp.exists():
					continue
				reference_data = self.api_get_id( wid )
				if not reference_data:
					reference_data = {}
				utils.write_json( reference_cached_fp , reference_data )

	# ---------- stats: single disk pass, populates self._index ----------

	def stats( self ):
		# 1. OpenAlex data we have for library papers
		zp = {}
		for fp in tqdm( list( self.storage_dir.glob( "*.json" ) ) , desc="Loading library" ):
			d = utils.read_json( fp ) or {}
			oid = d.get( "id" , "" )
			if isinstance( oid , str ) and oid.startswith( "https://openalex.org/" ):
				zp[ oid.rsplit( "/" , 1 )[ -1 ] ] = d

		# 2. Source of truth: Zotero snapshot
		snap = utils.zotero_simple_snapshot()
		lib_dois = { utils.normalize_doi( i[ "doi" ] ) for i in snap.values() if i.get( "doi" ) }
		lib_titles = { utils.openalex_normalize_title( i[ "title" ] ) for i in snap.values() if i.get( "title" ) }
		lib_wids = set( zp.keys() )

		# 3. Tally refs (skip library wids early)
		counts = Counter()
		for wid , p in tqdm( zp.items() , desc="Tallying refs" ):
			for r in p.get( "referenced_works" ) or []:
				rw = r.rsplit( "/" , 1 )[ -1 ]
				if rw in lib_wids:
					continue
				counts[ rw ] += 1

		# 4. Single pass: load each ref once, build row, stash haystack
		#    self._index keeps everything for downstream search sheets
		self._index = []
		rows , skipped = [] , 0
		for rw , n in tqdm( counts.most_common() , desc="Building rows" ):
			fp = self.references_dir / f"{rw}.json"
			meta = ( utils.read_json( fp ) or {} ) if fp.exists() else {}
			haystack = make_haystack( meta )

			rt = meta.get( "title" ) or meta.get( "display_name" )
			rt_norm = utils.openalex_normalize_title( rt ) if rt else None
			is_dup = bool( rt_norm and rt_norm in lib_titles )
			self._index.append( ( rw , meta , haystack , n , not is_dup ) )

			if is_dup:
				skipped += 1
				continue
			rows.append( build_row( rw , meta , n ) )

		# 5. Build base sheets
		sheets = [
			( "Top 1000 by Cites"   , HEADERS_ROW , rows[ :1000 ] ),
			( "Top 1000 by Recency" , HEADERS_ROW , sorted( rows , key=lambda r: r[ 2 ] or 0 , reverse=True )[ :1000 ] ),
		]

		utils.write_xlsx( self._xlsx_path , sheets )
		print( f"library={len(snap)} resolved={len(zp)} missing={len(rows)} (deduped {skipped}) -> {self._xlsx_path}" )

	# ---------- search sheets: pure in-memory, no disk re-read ----------

	def add_search_sheets( self , searches ):
		"""Append search-result sheets to missing.xlsx. Requires stats() to have run."""
		if not self._index:
			raise RuntimeError( "call stats() first — it builds the search index" )

		wb = load_workbook( self._xlsx_path )

		for name , predicate in searches:
			matched = []
			for wid , meta , hay , cite_count , included in tqdm( self._index , desc=f"Search: {name}" ):
				if not included:
					continue
				if not predicate( hay ):
					continue
				matched.append( build_row( wid , meta , cite_count ) )
			matched.sort( key=lambda r: r[ 2 ] or 0 , reverse=True )

			safe = safe_sheet_name( name )
			if safe in wb.sheetnames:
				del wb[ safe ]
			ws = wb.create_sheet( safe )
			ws.append( HEADERS_ROW )
			for r_idx , row in enumerate( matched[ :1000 ] , start=2 ):
				for c_idx , val in enumerate( row , start=1 ):
					cell = ws.cell( row=r_idx , column=c_idx )
					if isinstance( val , utils.Link ):
						cell.value = val.text
						cell.hyperlink = val.url
						cell.style = "Hyperlink"
					else:
						cell.value = val
			print( f"'{name}': {len(matched)} hits" )

		wb.save( self._xlsx_path )
		print( f"appended {len(searches)} search sheets -> {self._xlsx_path}" )


if __name__ == "__main__":
	x = OpenAlex()
	# x.update_cache()
	x.stats()
	x.add_search_sheets([
		( "Inner speech (any)" , any_of( "inner speech" , "imagined speech" , "covert speech" , "silent speech" , "subvocalized speech" , "inner monologue" , "imagined phonemes" , "silent communication" , "covert articulation" ) ),
		( "fMRI + inner speech" , combine_and( all_of( "fMRI" , "speech" ) , any_of( "inner" , "imagined" , "covert" ) ) ) ,
		# ( "MEG/EEG + imagined" , combine_and( any_of( "MEG" , "EEG" ) , any_of( "imagined" , "covert" ) ) ),
	])