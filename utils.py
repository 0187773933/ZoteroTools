import os
import re
import json
import csv
import tempfile
import shutil
import unicodedata
from pprint import pprint
from pathlib import Path
from tqdm import tqdm
from urllib.parse import urlparse, unquote
import sqlite3
import base64
from openpyxl import Workbook
from openpyxl.styles import Font

def write_json( file_path , python_object ):
	with open( file_path , 'w', encoding='utf-8' ) as f:
		json.dump( python_object , f , ensure_ascii=False , indent=4 )

def read_json( file_path ):
	with open( file_path ) as f:
		return json.load( f )

def normalize_doi(value: str) -> str:
	if not value:
		return value
	v = value.strip()
	v = v.replace("https://doi.org/", "").replace("http://doi.org/", "")
	return v.strip()

def normalize_title(s: str) -> str:
	if not s:
		return ""

	# unicode normalization
	s = unicodedata.normalize("NFKD", s)

	# lowercase
	s = s.lower()

	# replace punctuation with space
	s = re.sub(r"[^a-z0-9]+", " ", s)

	# collapse whitespace
	s = " ".join(s.split())

	return s.strip()

def write_csv( csv_path , headers , rows ):
	with open( csv_path , "w" , newline="" ) as f:
		writer = csv.DictWriter(
			f , fieldnames=headers )
		writer.writeheader()
		for row in rows:
			writer.writerow( row )

def base64_encode( message ):
	try:
		message_bytes = message.encode( 'utf-8' )
		base64_bytes = base64.b64encode( message_bytes )
		base64_message = base64_bytes.decode( 'utf-8' )
		return base64_message
	except Exception as e:
		print( e )
		return False

def base64_decode( base64_message ):
	try:
		base64_bytes = base64_message.encode( 'utf-8' )
		message_bytes = base64.b64decode(base64_bytes)
		message = message_bytes.decode( 'utf-8' )
		return message
	except Exception as e:
		print( e )
		return False

def doi_fp( doi ):
	return re.sub( r'[^a-zA-Z0-9._-]', '_' , doi )

def _candidates_common(home):
	return [
		home / "Zotero" / "zotero.sqlite",
		home / "ZoteroBeta" / "zotero.sqlite",
		home / "Zotero Beta" / "zotero.sqlite",
		home / "Library" / "Application Support" / "Zotero" / "zotero.sqlite",
		home / "Library" / "Application Support" / "ZoteroBeta" / "zotero.sqlite",
		home / "Library" / "Application Support" / "Zotero Beta" / "zotero.sqlite",
	]

def _bounded_find_sqlite(home):
	roots = [
		home / "Zotero",
		home / "Library" / "Application Support" / "Zotero",
		home / "Library" / "Application Support",
	]
	roots = [r for r in roots if r.exists()]

	best: Optional[Tuple[float, Path]] = None  # (mtime, path)

	for root in roots:
		for pat in ("zotero.sqlite", "**/zotero.sqlite"):
			try:
				for p in root.glob(pat):
					if p.name != "zotero.sqlite":
						continue
					try:
						st = p.stat()
					except OSError:
						continue
					if best is None or st.st_mtime > best[0]:
						best = (st.st_mtime, p)
			except Exception:
				continue

	return best[1] if best else None

def resolve_zotero_db_path( cli_db ):
	# 1) CLI
	if cli_db:
		p = Path(cli_db).expanduser()
		if p.exists():
			return p
		raise SystemExit(f"--db path does not exist: {p}")

	# 2) ENV
	env = os.environ.get("ZOTERO_DB", "").strip()
	if env:
		p = Path(env).expanduser()
		if p.exists():
			return p
		raise SystemExit(f"ZOTERO_DB path does not exist: {p}")

	home = Path.home()

	# 3) Common locations
	for p in _candidates_common(home):
		if p.exists():
			return p

	# 4) Bounded search
	p = _bounded_find_sqlite(home)
	if p and p.exists():
		return p

	raise SystemExit(
		"Could not find zotero.sqlite automatically.\n"
		"Provide --db /path/to/zotero.sqlite or set ZOTERO_DB=/path/to/zotero.sqlite"
	)

ZOTERO_DB = resolve_zotero_db_path( None )
ZOTERO_STORAGE = ZOTERO_DB.parent / "storage"
def zotero_open_snapshot():
	tmpdir = Path(tempfile.mkdtemp( prefix="zotero_db_" ) )
	tmpdb = tmpdir / "zotero.sqlite"
	shutil.copy2( ZOTERO_DB , tmpdb )
	conn = sqlite3.connect( tmpdb )
	conn.row_factory = sqlite3.Row
	return conn

def resolve_zotero_path( path , key ):
	if not path:
		return None

	if path.startswith("storage:"):
		rel = path.replace("storage:", "", 1)
		return ZOTERO_STORAGE / key / rel

	if path.startswith("file:"):
		from urllib.parse import urlparse, unquote
		return Path(unquote(urlparse(path).path))

	return None

def zotero_take_snapshot():
	conn = zotero_open_snapshot()
	c = conn.cursor()

	# --------------------------------------------------
	# 1) BASE ITEMS: ONLY "REAL" BIB ITEMS (exclude attachments/notes/annotations)
	# --------------------------------------------------
	# Zotero's UI count (~681) corresponds to bibliographic items, not the raw items table.
	EXCLUDE_TYPES = ( "attachment" , "note" , "annotation" )

	papers = {}

	for row in c.execute("""
		SELECT items.itemID, items.key, itemTypes.typeName
		FROM items
		JOIN itemTypes ON itemTypes.itemTypeID = items.itemTypeID
		LEFT JOIN deletedItems ON deletedItems.itemID = items.itemID
		WHERE deletedItems.itemID IS NULL
		  AND itemTypes.typeName NOT IN ('attachment','note','annotation')
	"""):
		itemID = row["itemID"]
		papers[itemID] = {
			"itemID": itemID,
			"key": row["key"],
			"type": row["typeName"],
			"doi": None,
			"attachments": [],
			"meta": {},
			"creators": [],
			"tags": [],
			"collections": []
		}

	# --------------------------------------------------
	# 2) METADATA (title, DOI, journal, year, etc)
	# --------------------------------------------------
	for row in c.execute("""
		SELECT itemData.itemID, fields.fieldName, itemDataValues.value
		FROM itemData
		JOIN fields ON fields.fieldID = itemData.fieldID
		JOIN itemDataValues ON itemDataValues.valueID = itemData.valueID
	"""):
		itemID = row["itemID"]
		if itemID not in papers:
			continue

		field = row["fieldName"]
		value = row["value"]

		papers[itemID]["meta"][field] = value

		if field == "DOI" and value:
			papers[itemID]["doi"] = normalize_doi(value)

	# --------------------------------------------------
	# 3) CREATORS (authors/editors)
	# --------------------------------------------------
	for row in c.execute("""
		SELECT itemCreators.itemID,
			   creators.firstName,
			   creators.lastName,
			   creatorTypes.creatorType
		FROM itemCreators
		JOIN creators ON creators.creatorID = itemCreators.creatorID
		JOIN creatorTypes ON creatorTypes.creatorTypeID = itemCreators.creatorTypeID
		ORDER BY itemCreators.itemID, itemCreators.orderIndex
	"""):
		itemID = row["itemID"]
		if itemID not in papers:
			continue

		papers[itemID]["creators"].append({
			"type": row["creatorType"],
			"first": row["firstName"],
			"last": row["lastName"]
		})

	# --------------------------------------------------
	# 4) ALL ATTACHMENTS (child items) grouped onto their parent bib item
	# --------------------------------------------------
	for row in c.execute("""
		SELECT itemAttachments.parentItemID AS parentID,
			   itemAttachments.itemID       AS attachItemID,
			   items.key                    AS attachKey,
			   itemAttachments.path         AS path,
			   itemAttachments.contentType  AS contentType,
			   itemAttachments.linkMode     AS linkMode
		FROM itemAttachments
		JOIN items ON items.itemID = itemAttachments.itemID
	"""):

		parentID = row["parentID"]

		# if missing parent, create placeholder
		if parentID not in papers:
			papers[parentID] = {
				"itemID": parentID,
				"key": None,
				"type": "unknown",
				"doi": None,
				"attachments": [],
				"meta": {},
				"creators": [],
				"tags": [],
				"collections": []
			}

		path = row["path"]
		attachKey = row["attachKey"]

		# file_path = None

		# if path and path.startswith("storage:"):
		# 	candidate = ZOTERO_STORAGE / attachKey / path.replace("storage:", "")
		# 	if candidate.exists():
		# 		file_path = candidate

		if path:
			if path.startswith("storage:"):
				rel = path.replace("storage:", "", 1)
				file_path = ZOTERO_STORAGE / attachKey / rel

			elif path.startswith("file:"):
				file_path = Path(unquote(urlparse(path).path))

		papers[parentID]["attachments"].append({
			"key": attachKey,
			"parent_id": parentID ,
			"contentType": row["contentType"],
			"linkMode": row["linkMode"],
			"path": path,
			"abs_path": str(file_path) if file_path else None
		})

	# --------------------------------------------------
	# 5) TAGS (only for base bib items)
	# --------------------------------------------------
	for row in c.execute("""
		SELECT itemTags.itemID, tags.name
		FROM itemTags
		JOIN tags ON tags.tagID = itemTags.tagID
	"""):
		itemID = row["itemID"]
		if itemID not in papers:
			continue
		papers[itemID]["tags"].append(row["name"])

	# --------------------------------------------------
	# 6) COLLECTIONS (only for base bib items)
	# --------------------------------------------------
	for row in c.execute("""
		SELECT collectionItems.itemID, collections.collectionName
		FROM collectionItems
		JOIN collections ON collections.collectionID = collectionItems.collectionID
	"""):
		itemID = row["itemID"]
		if itemID not in papers:
			continue
		papers[itemID]["collections"].append(row["collectionName"])

	conn.close()

	# Sort tag/collection lists for stability
	for item in papers.values():
		item["tags"] = sorted(set(item["tags"]))
		item["collections"] = sorted(set(item["collections"]))

	# Return keyed by Zotero key (one per bib item)
	return {item["key"]: item for item in papers.values()}


def zotero_simple_snapshot():
	snapshot = zotero_take_snapshot()
	papers = {}
	for key in snapshot:
		item = snapshot[ key ]
		item_id = str( item.get( "itemID" ) )
		doi = item.get( "doi" ) or item.get( "meta" ).get( "DOI" )
		title = item.get( "title" ) or item.get( "meta" ).get( "title" )
		url = item.get( "meta" ).get( "url" )
		date = item.get( "meta" ).get( "date" )
		pdf_paths = [
			p.get( "abs_path" )
			for p in item.get( "attachments" , [] )
			if p.get( "abs_path" ) and p.get( "abs_path" ).lower().endswith( ".pdf" )
		]
		paper = {
			"doi": doi ,
			"id": item_id ,
			"title": title ,
			"url": url ,
			"date": date ,
			"pdfs": pdf_paths
		}
		papers[ key ] = paper
	return papers


def openalex_normalize_title( title ):
	if not title:
		return ""

	title = unicodedata.normalize( "NFKD" , str( title ) )
	title = title.replace( "|" , " " )
	title = title.replace( "“" , '"' ).replace( "”" , '"' )
	title = title.replace( "‘" , "'" ).replace( "’" , "'" )

	# Remove chars that OpenAlex search parser may treat as syntax/operators.
	title = re.sub( r"[|(){}\[\]^~?:\\/]" , " " , title )

	# Collapse whitespace.
	title = re.sub( r"\s+" , " " , title ).strip()

	return title

class Link:
	__slots__ = ( "text" , "url" )
	def __init__( self , text , url ):
		self.text , self.url = text , url
def write_xlsx( filepath , sheets ):
	wb = Workbook()
	wb.remove( wb.active )
	bold = Font( bold=True )
	for name , headers , rows in sheets:
		ws = wb.create_sheet( name )
		ws.append( list( headers ) )
		for c in ws[ 1 ]:
			c.font = bold
		for row in tqdm( list( rows ) , desc=f"Writing {name}" ):
			row = list( row )
			ws.append([ ( c.text if isinstance( c , Link ) else c ) for c in row ])
			for col_idx , c in enumerate( row , 1 ):
				if isinstance( c , Link ) and c.url:
					cell = ws.cell( row=ws.max_row , column=col_idx )
					cell.hyperlink = c.url
					cell.style = "Hyperlink"
	wb.save( filepath )